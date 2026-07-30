#!/usr/bin/env python3
"""
derive_subsumption_gold.py — Subsumptions-Gold aus dem Äquivalenz-Mapping ableiten
===================================================================================
Inputs : vdi-ebay.xlsx            (Hand-Mapping: code -> eBay-Kategorie, implizit '=')
         VDI_5_6.xlsx             (Source-Hierarchie: Gold_Baugruppen/-Relationen/-Kategorien)
         vdi_position_pairs_chat.tsv (16 Positionsklassen v{ID}_{Code})
         ebay_kfz_tree.csv        (Target-Hierarchie)
Outputs: gold_relations_karosserie.tsv   (Seeds + verfeinerte Relationen + transitive Hülle)
         gold_subsumption_findings.md    (alle Eingriffe, Konflikte, Ausschlüsse — vetobar)

DOKUMENTIERTE KORREKTUR (A. Markic, 17.07.2026): VDI4081_10675 (Basis) Ziel 262162 -> 262166
(Erfassungsfehler, eine Endziffer; alle sechs Positionszeilen derselben Quelle treffen 262166).

RELATIONSKONVENTION (source-relativ, wie Pipeline/OAEI-Referenzen):
  v < e : VDI-Konzept ist Unterklasse der eBay-Kategorie
  v > e : VDI-Konzept ist Oberklasse der eBay-Kategorie
  v = e : Äquivalenz

VERFEINERUNGS-HEURISTIKEN (Seeds; jede Anwendung wird im Findings-Report gelistet):
  H1  1:1-Zuordnung ohne Widerspruch                          -> '=' bleibt
  H2  Positions-/Varianten-Zeile mit gleichem Ziel wie Basis  -> Zeile entfällt als Seed
      (Relation entsteht in der Hülle: Kind ⊑ Basis = Ziel ⇒ Kind < Ziel)
  H3  Mehrere NICHT hierarchisch verwandte Quellkonzepte auf  -> alle '<'
      dasselbe Ziel (Sammelknoten-Signatur)
  H4  Ziel-Label beginnt mit 'Sonstige'/'Weitere'             -> '<' (Catch-all-Bucket)
  H5  Positionsklasse auf positionsspezifisches Sammelziel    -> '<' (enger als Knoten,
      der links+rechts bzw. mehrere Bauteilarten umfasst)
  H6  Doppelziel-Basis (Kinder zeigen auf anderes Ziel)       -> '>' zum eigenen Ziel
      (Basis deckt mehr ab als die positionsbeschränkte Zielkategorie)
  H7  Hierarchisch verwandte Quellen auf demselben Ziel        -> nur das oberste Konzept
      behält '='; alle Nachfahren '<' (verhindert Ketten-Äquivalenzen)
  Prioritäten je Seed: H2 > H6 > H5 > H3/H4 > H1. H3/H4 ändern nie ein durch H6 gesetztes '>'.

TRANSITIVE HÜLLE (nur subClassOf; partOf wird bewusst NICHT traversiert):
  (v,e,=): v<anc_T(e) | v>desc_T(e) | desc_S(v)<e | anc_S(v)>e | desc_S(v)<anc_T(e) | anc_S(v)>desc_T(e)
  (v,e,<): v<anc_T(e) | desc_S(v)<e | desc_S(v)<anc_T(e)
  (v,e,>): v>desc_T(e) | anc_S(v)>e | anc_S(v)>desc_T(e)
  owl:Thing ausgeschlossen; Hauptgruppen-Klassen enthalten, aber source_level='hauptgruppe'.
  Konflikte (gleiches Paar mit < UND >) werden NICHT emittiert, sondern rapportiert.
"""
import csv
import re
from collections import defaultdict, Counter
from openpyxl import load_workbook

# Repo-relative Pfade (lokale Anpassung; Original-Lieferpfade siehe subsumption_gold.zip).
from pathlib import Path
_ROOT = Path(__file__).resolve().parent.parent.parent   # workspace/olala
F_MAP = str(_ROOT / 'vdi-ebay.xlsx')
F_VDI = str(_ROOT / 'goldstandard_ebay' / 'vdi' / 'VDI_5_6.xlsx')
F_POS = str(_ROOT / 'goldstandard_ebay' / 'vdi_position_pairs.tsv')
F_EBAY = str(_ROOT / 'goldstandard_ebay' / 'ebay_kfz_tree.csv')
OUT_TSV = str(Path(__file__).resolve().parent / 'gold_relations_karosserie.tsv')
OUT_MD = str(Path(__file__).resolve().parent / 'gold_subsumption_findings.md')

CORRECTIONS = {('VDI4081_10675', None): 262166}   # dokumentierter Erfassungsfehler (statt 262162)

# DOKUMENTIERTE ADJUDIKATION (A. Markic, 2026-07-17): vier 1:1-'='-Seeds, deren
# Ziel erkennbar breiter ist, werden auf '<' gesetzt — strenge Mengensemantik,
# konsistent mit der H4-Behandlung der Catch-all-Buckets ("bestpassende
# Sammelkategorie" gilt als 'feiner als', nicht 'äquivalent'). Die Mechanik
# (H1-H7) konnte sie nicht kippen, weil sie im Mapping 1:1 sind; Fachentscheid.
RELATION_OVERRIDES = {
    'v10356': '<',   # Türscheibe            < Autoglas (alle Fahrzeugscheiben)
    'v11535': '<',   # Kupplungspedal        < Pedale, Pedalkappen & Fußstützen
    'v10065': '<',   # Kennzeichenverstärkung < Nummernschilder & -rahmen
    'v11878': '<',   # Radarsensor           < Kameras, Überwachungsgeräte & Sensorsätze
}
HGN = {1: 'Unterboden', 2: 'Motorraum', 3: 'Karosserie', 4: 'Innenraum', 5: 'Sonstige'}

# ---------- Target-Hierarchie (eBay) ----------
e_label, e_parents = {}, defaultdict(set)
with open(F_EBAY, encoding='utf-8') as f:
    for row in csv.DictReader(f):
        i = int(row['id']); e_label[i] = row['label']
        if row['parent_id']: e_parents[i].add(int(row['parent_id']))

def anc_T(i):
    out, stack = {}, [(p, 1) for p in e_parents.get(i, ())]
    while stack:
        n, d = stack.pop()
        if n in out and out[n] <= d: continue
        out[n] = d
        stack += [(p, d + 1) for p in e_parents.get(n, ())]
    return out
e_children = defaultdict(set)
for c, ps in e_parents.items():
    for p in ps: e_children[p].add(c)
def desc_T(i):
    out, stack = {}, [(c, 1) for c in e_children.get(i, ())]
    while stack:
        n, d = stack.pop()
        if n in out and out[n] <= d: continue
        out[n] = d
        stack += [(c, d + 1) for c in e_children.get(n, ())]
    return out

# ---------- Source-Hierarchie (VDI 5.6 + Positionsklassen) ----------
wb = load_workbook(F_VDI, data_only=True)
v_label, v_hg, v_level = {}, {}, {}
for r in wb['Gold_Baugruppen'].iter_rows(min_row=2, values_only=True):
    if r[0] is None: continue
    i = int(r[0]); v_label[f'v{i}'] = str(r[1]); v_hg[f'v{i}'] = str(r[2]); v_level[f'v{i}'] = 'baugruppe'
for r in wb['Gold_Kategorien'].iter_rows(min_row=2, values_only=True):
    if r[0] is None: continue
    v_label[f'v{int(r[0])}'] = str(r[1]); v_level[f'v{int(r[0])}'] = 'kategorie'
for n, name in HGN.items():
    v_label[f'hg{n}'] = name; v_level[f'hg{n}'] = 'hauptgruppe'
v_parents = defaultdict(set)
sub_children_ids = set()
for r in wb['Gold_Relationen'].iter_rows(min_row=2, values_only=True):
    if r[0] is None or str(r[2]) != 'subClassOf': continue
    c, p = int(r[0]), int(r[3])
    sub_children_ids.add(c)
    v_parents[f'v{c}'].add(f'hg{p}' if p in HGN else f'v{p}')
for key in list(v_label):
    if key.startswith('v') and v_level[key] == 'baugruppe':
        i = int(key[1:])
        if i not in sub_children_ids:
            v_parents[key].add(f"hg{v_hg[key].strip()[0]}")     # HG-Verankerung wie in vdi2owl
with open(F_POS, encoding='utf-8') as f:
    for line in f:
        if not line.strip(): continue
        i, code = line.split()
        key = f'v{i}_{code}'
        v_parents[key].add(f'v{i}')
        v_label[key] = None      # Label später aus Basis + Code (nur informativ)
        v_level[key] = 'position'
POS_DE = {1:'links',2:'rechts',16:'vorne',17:'vorne links',18:'vorne rechts',65:'hinten links',66:'hinten rechts'}
for key in v_label:
    if v_level.get(key) == 'position':
        base, code = key.rsplit('_', 1)
        v_label[key] = f"{v_label[base]}, {POS_DE.get(int(code), code)}"

v_children = defaultdict(set)
for c, ps in v_parents.items():
    for p in ps: v_children[p].add(c)
def anc_S(k):
    out, stack = {}, [(p, 1) for p in v_parents.get(k, ())]
    while stack:
        n, d = stack.pop()
        if n in out and out[n] <= d: continue
        out[n] = d
        stack += [(p, d + 1) for p in v_parents.get(n, ())]
    return out
def desc_S(k):
    out, stack = {}, [(c, 1) for c in v_children.get(k, ())]
    while stack:
        n, d = stack.pop()
        if n in out and out[n] <= d: continue
        out[n] = d
        stack += [(c, d + 1) for c in v_children.get(n, ())]
    return out

# ---------- Mapping einlesen, normalisieren, Scope ----------
wbm = load_workbook(F_MAP, data_only=True)
rows = list(wbm['in'].iter_rows(min_row=2, values_only=True))
def ebay_id(v):
    for cand in (round(v * 1000), round(v)):
        if cand in e_label: return cand
    return None
raw, excl = [], Counter()
bad_targets = []
for r in rows:
    code, target_raw, state = r[2], r[4], r[7]
    if state != 'ACTIVE': excl['INACTIVE'] += 1; continue
    m = re.fullmatch(r'VDI4081_(\d+)(?:_(\d+))?', str(code).strip())
    if not m: excl['Code unlesbar'] += 1; bad_targets.append((code, 'Code')); continue
    base, suf = int(m.group(1)), (int(m.group(2)) if m.group(2) else None)
    t = CORRECTIONS.get((f'VDI4081_{base}', suf))
    if t is None: t = ebay_id(float(target_raw))
    if t is None: excl['Ziel nicht im eBay-Baum'] += 1; bad_targets.append((code, target_raw)); continue
    if f'v{base}' not in v_label: excl['außerhalb Karosserie-Gold'] += 1; continue
    raw.append((base, suf, t))
n_corr = sum(1 for (b, s, t) in raw if CORRECTIONS.get((f'VDI4081_{b}', s)) == t)

# ---------- Kollaps + Seeds ----------
by_base = defaultdict(list)
for base, suf, t in raw: by_base[base].append((suf, t))
POS_KEYS = {k for k in v_label if v_level.get(k) == 'position'}
seeds, findings = [], defaultdict(list)
for base, entries in sorted(by_base.items()):
    targets = {t for _, t in entries}
    base_t = next((t for s, t in entries if s is None), None)
    if len(targets) == 1:
        t = targets.pop()
        seeds.append([f'v{base}', t, '=', 'Seed'])
        n_dup = len(entries) - 1
        if n_dup: findings['H2'].append(f"v{base}: {n_dup} Suffixzeile(n) zielgleich kollabiert -> Hülle")
    else:
        for suf, t in entries:
            if suf is None:
                seeds.append([f'v{base}', t, '>', 'Seed+H6'])
                findings['H6'].append(f"v{base} '{v_label[f'v{base}']}' > e{t} '{e_label[t]}' (Doppelziel-Basis)")
            else:
                key = f'v{base}_{suf}'
                if key in POS_KEYS:
                    if t == base_t:
                        findings['H2'].append(f"{key}: zielgleich mit Basis -> Hülle"); continue
                    seeds.append([key, t, '<', 'Seed+H5'])
                    findings['H5'].append(f"{key} '{v_label[key]}' < e{t} '{e_label[t]}'")
                else:
                    if t == base_t:
                        findings['H2'].append(f"v{base}_{suf}: zielgleich mit Basis (keine Positionsklasse) -> Hülle"); continue
                    excl['Suffixzeile ohne Positionsklasse, Ziel abweichend'] += 1
                    bad_targets.append((f'VDI4081_{base}_{suf}', t))

# H4, dann H7 (verwandte Gruppen: nur Spitze behält '='), dann H3 (unverwandte Spitzen)
for s in seeds:
    if s[2] != '=': continue
    lbl = e_label[s[1]]
    if lbl.startswith('Sonstige') or lbl.startswith('Weitere'):
        s[2] = '<'; s[3] += '+H4'; findings['H4'].append(f"{s[0]} '{v_label[s[0]]}' < e{s[1]} '{lbl}'")
tgt_groups = defaultdict(list)
for s in seeds:
    if s[2] == '=': tgt_groups[s[1]].append(s)
for t, members in tgt_groups.items():
    keys = {m[0] for m in members}
    tops = []
    for m in members:
        if keys & set(anc_S(m[0])):
            m[2] = '<'; m[3] += '+H7'
            findings['H7'].append(f"{m[0]} '{v_label[m[0]]}' < e{t} '{e_label[t]}' (Vorfahre in derselben Zielgruppe behält '=')")
        else:
            tops.append(m)
    if len(tops) > 1:
        for m in tops:
            m[2] = '<'; m[3] += '+H3'
            findings['H3'].append(f"{m[0]} '{v_label[m[0]]}' < e{t} '{e_label[t]}' (teilt Ziel mit {len(tops)-1} unverwandten Spitzen)")

# ADJ: adjudizierte Relations-Overrides (nach H1-H7, vor der Hülle, damit die
# geflippten Seeds nur noch '<'-propagieren).
for s in seeds:
    want = RELATION_OVERRIDES.get(s[0])
    if want and s[2] == '=' :
        s[2] = want; s[3] += '+ADJ'
        findings['ADJ'].append(f"{s[0]} '{v_label[s[0]]}' {want} e{s[1]} '{e_label[s[1]]}' (Adjudikation A. Markic 2026-07-17)")

# ---------- Transitive Hülle ----------
best = {}   # (v,e) -> [rel, hops_s, hops_t, rules]
def emit(v, e, rel, hs, ht, rule):
    k = (v, e)
    cur = best.get(k)
    if cur and cur[0] != rel:
        conflicts.append((v, e, cur[0], rel, cur[3], rule)); return
    if cur is None or (hs + ht) < (cur[1] + cur[2]):
        best[k] = [rel, hs, ht, rule if cur is None else f"{cur[3]}|{rule}"]
    elif rule not in cur[3]:
        cur[3] += f"|{rule}"
conflicts = []
for v, e, rel, srule in seeds:
    emit(v, e, rel, 0, 0, srule)
    aT, dT, aS, dS = anc_T(e), desc_T(e), anc_S(v), desc_S(v)
    if rel in ('=', '<'):
        for a, ht in aT.items(): emit(v, a, '<', 0, ht, f'Hülle<-{srule}')
        for s, hs in dS.items(): emit(s, e, '<', hs, 0, f'Hülle<-{srule}')
        for s, hs in dS.items():
            for a, ht in aT.items(): emit(s, a, '<', hs, ht, f'Hülle<-{srule}')
    if rel in ('=', '>'):
        for d, ht in dT.items(): emit(v, d, '>', 0, ht, f'Hülle<-{srule}')
        for s, hs in aS.items(): emit(s, e, '>', hs, 0, f'Hülle<-{srule}')
        for s, hs in aS.items():
            for d, ht in dT.items(): emit(s, d, '>', hs, ht, f'Hülle<-{srule}')

# ---------- Ausgabe ----------
seed_keys = {(v, e) for v, e, _, _ in seeds}
with open(OUT_TSV, 'w', encoding='utf-8', newline='') as f:
    w = csv.writer(f, delimiter='\t')
    w.writerow(['source', 'source_label', 'source_level', 'target', 'target_label',
                'relation', 'tier', 'hops_source', 'hops_target', 'rules'])
    for (v, e), (rel, hs, ht, rules) in sorted(best.items(), key=lambda x: (x[1][1] + x[1][2], x[0])):
        tier = 'seed' if (v, e) in seed_keys else 'derived'
        w.writerow([v, v_label.get(v, ''), v_level.get(v, ''), f'e{e}', e_label[e], rel, tier, hs, ht, rules])

tiers = Counter('seed' if k in seed_keys else 'derived' for k in best)
rels = Counter(val[0] for val in best.values())
seed_rels = Counter(s[2] for s in seeds)
with open(OUT_MD, 'w', encoding='utf-8') as f:
    f.write("# Subsumptions-Gold — Ableitungsreport\n\n")
    f.write(f"Mapping-Zeilen gesamt: {len(rows)} | in Scope übernommen: {len(raw)} | Korrektur angewendet: {n_corr} (10675->262166)\n\n")
    f.write(f"Ausschlüsse: {dict(excl)}\n\n")
    f.write(f"Seeds: {len(seeds)} ({dict(seed_rels)})\nGesamtpaare: {len(best)} ({dict(tiers)}; Relationen {dict(rels)})\nKonflikte: {len(conflicts)}\n\n")
    for h in ('H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'ADJ'):
        f.write(f"## {h} ({len(findings[h])})\n")
        for line in findings[h][:400]: f.write(f"- {line}\n")
        f.write("\n")
    f.write(f"## Konflikte ({len(conflicts)})\n")
    for c in conflicts[:200]: f.write(f"- {c}\n")
    f.write(f"\n## Nicht auflösbare Zeilen ({len(bad_targets)})\n")
    for b in bad_targets[:100]: f.write(f"- {b}\n")

print(f"Mapping: {len(rows)} Zeilen | ACTIVE in Scope: {len(raw)} | Ausschlüsse: {dict(excl)}")
print(f"Seeds: {len(seeds)} {dict(seed_rels)} | Paare gesamt: {len(best)} {dict(rels)} | Tiers: {dict(tiers)}")
print(f"Heuristik-Anwendungen: " + ", ".join(f"{h}: {len(findings[h])}" for h in ('H2','H3','H4','H5','H6','H7')))
print(f"Konflikte: {len(conflicts)}")
