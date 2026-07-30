#!/usr/bin/env python3
"""
vdi2owl.py — Wiederverwendbarer Konverter: VDI-Goldstandard (.xlsx) -> Source-Ontologie (.owl)
==============================================================================================
Erzeugt aus einer Gold-Datei im FINAL-Format (Blätter: Gold_Baugruppen, Gold_Relationen,
Gold_Kategorien) die VDI-Source-Ontologie. Funktioniert unverändert weiter, wenn der
Goldstandard später erweitert wird (mehr Zeilen, weitere Schwerpunkte, neue Kategorien,
neue Anker) — alle Inhalte werden aus der Datei gelesen, nichts ist hartkodiert.

Benötigt: pip install openpyxl rdflib

Aufrufe:
  python3 vdi2owl.py VDI_GoldStandard_FINAL.xlsx
  python3 vdi2owl.py FINAL.xlsx -o vdi_source.owl --stammliste VDI_GoldStandard_ExperteB.xlsx
  python3 vdi2owl.py FINAL.xlsx --partof direct --base-uri "http://example.org/vdi#"

Argumente:
  input                 Gold-Datei (.xlsx) im FINAL-Format
  -o / --output         Ziel-OWL (Default: <input>.owl neben der Eingabe)
  --stammliste PATH     Beliebige Template-Datei mit Blatt 'Baugruppen' (z.B. Experten- oder
                        Originaltemplate). NUR nötig, wenn Gold_Relationen auf Baugruppen-IDs
                        zeigen, die weder in Gold_Baugruppen noch in Gold_Kategorien stehen
                        (Anker wie 10159 Tank) — von dort kommen Label und Hauptgruppe.
  --base-uri            URI-Präfix der Klassen (Default: http://vdi.de/kfz#)
  --ontology-uri        Ontologie-URI (Default: base-uri ohne '#')
  --partof {restriction,direct}
                        'restriction' (Default) = OAEI-Anatomy-Muster:
                            Child rdfs:subClassOf [ owl:Restriction onProperty part_of
                                                    owl:someValuesFrom Parent ]
                        'direct' = einfache Tripel  Child part_of Parent

Modellierung (identisch zur Erstauslieferung vom 15.07.2026):
  - subClassOf -> rdfs:subClassOf; Hauptgruppen (1-5) als Top-Klassen hg1..hg5 unter owl:Thing
  - Baugruppen ohne expliziten sub-Parent werden an ihrer Hauptgruppen-Klasse verankert
  - Hauptgruppe_Final zusätzlich als Annotation vdi:hauptgruppe an jeder Baugruppen-Klasse
  - URI-Schema: v{ID} für Baugruppen/Kategorien, hg{N} für Hauptgruppen
Validierung: Blatt-/Spaltenprüfung, unbekannte Relationen, nicht auflösbare Parent-IDs,
  Kategorien ohne sub-Parent, Zyklen je Relation, Vollständigkeits- und Kantenzählung.

Einbauposition-Ebene (optional, additiv):
  --vdi-original VDI_5_5_6.xlsx   Original-VDI-Dokument; Quelle der Einbauort-Legende
                                  (Blatt 'Einbauort': Position-Code-Tabelle + Bitmasken-
                                  Atome LEFT=1, FRONT=16, REAR=64, BOTTOM=512, ...)
  --positionen SPEC_ODER_DATEI    Welche Positionsklassen erzeugt werden:
                                  Spec-String  "10610:16,17,18;10675:65,66"
                                  oder TSV/CSV-Datei mit Spalten id, code (eine Zeile je Paar)
  Erzeugt je Paar eine Klasse v{ID}_{Code} rdfs:subClassOf v{ID} mit Label
  "{Baugruppenlabel}, {Positionslabel}" (z.B. "Kotflügel, hinten links") und Annotation
  vdi:einbauposition. Codes ohne Legendeneintrag werden über die Bitmasken-Atome
  dekomponiert (528 = 512+16 -> "unten vorne"); nicht dekomponierbare Codes -> Fehler.
"""
import argparse
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from rdflib import Graph, Namespace, URIRef, Literal, BNode, RDF, RDFS, OWL

HGN = {1: "Unterboden", 2: "Motorraum", 3: "Karosserie", 4: "Innenraum", 5: "Sonstige"}
LANG = "de"


def die(msg):
    print(f"FEHLER: {msg}", file=sys.stderr)
    sys.exit(1)


def hg_num(s):
    s = str(s).strip()
    if not s or not s[0].isdigit() or int(s[0]) not in HGN:
        return None
    return int(s[0])


def read_gold(path):
    wb = load_workbook(path, data_only=True)
    for sheet in ("Gold_Baugruppen", "Gold_Relationen", "Gold_Kategorien"):
        if sheet not in wb.sheetnames:
            die(f"Blatt '{sheet}' fehlt in {path} — ist das eine Datei im FINAL-Format?")
    bau, rels, cats = {}, [], {}
    for r in wb["Gold_Baugruppen"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        i, label, hg = int(r[0]), str(r[1]).strip(), str(r[2]).strip()
        if hg_num(hg) is None:
            die(f"Gold_Baugruppen ID {i}: Hauptgruppe_Final unlesbar: {hg!r}")
        bau[i] = (label, hg)
    for r in wb["Gold_Relationen"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        rel = str(r[2]).strip()
        if rel not in ("subClassOf", "partOf"):
            die(f"Gold_Relationen: unbekannte Relation {rel!r} bei child {r[0]}")
        rels.append((int(r[0]), rel, int(r[3])))
    for r in wb["Gold_Kategorien"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        cats[int(r[0])] = str(r[1]).strip()
    if not bau:
        die("Gold_Baugruppen ist leer.")
    return bau, rels, cats


def read_stammliste(path):
    """Label/Hauptgruppe für Anker-IDs. Erkennt zwei Formate:
    - Template-Blatt 'Baugruppen'            (ID=Sp.1, Bezeichnung=Sp.2, Hauptgruppe_Final=Sp.6)
    - Original-Blatt 'Bezeichnungen_Synonyme' (ID=Sp.1, Produkthauptgruppe=Sp.3, Bezeichnung=Sp.17)
    """
    wb = load_workbook(path, data_only=True)
    out = {}
    if "Baugruppen" in wb.sheetnames:
        for r in wb["Baugruppen"].iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            out[int(r[0])] = (str(r[1]).strip(), str(r[5]).strip() if r[5] else None)
    elif "Bezeichnungen_Synonyme" in wb.sheetnames:
        for r in wb["Bezeichnungen_Synonyme"].iter_rows(min_row=2, values_only=True):
            if r[0] is None:
                continue
            out[int(r[0])] = (str(r[16]).strip(), str(r[2]).strip() if r[2] is not None else None)
    else:
        die(f"--stammliste {path}: weder Blatt 'Baugruppen' noch 'Bezeichnungen_Synonyme' gefunden.")
    if not out:
        die(f"--stammliste {path}: keine Zeilen gelesen.")
    return out


def cycle_check(edges, relname):
    graph = defaultdict(list)
    for c, p in edges:
        graph[c].append(p)
    color, cycles = defaultdict(int), []

    def dfs(n, stack):
        color[n] = 1
        stack.append(n)
        for m in graph.get(n, []):
            if color[m] == 1:
                cycles.append(stack[stack.index(m):] + [m])
            elif color[m] == 0:
                dfs(m, stack)
        stack.pop()
        color[n] = 2

    for n in list(graph):
        if color[n] == 0:
            dfs(n, [])
    if cycles:
        die(f"{relname}-Zyklen gefunden: {cycles}")


def read_einbauort_legende(path):
    """Liest Blatt 'Einbauort': explizite Position-Code-Tabelle (Code -> de-Label)."""
    wb = load_workbook(path, data_only=True)
    if "Einbauort" not in wb.sheetnames:
        die(f"--vdi-original {path}: Blatt 'Einbauort' fehlt.")
    leg = {}
    for r in wb["Einbauort"].iter_rows(values_only=True):
        if r and isinstance(r[0], int) and isinstance(r[1], str) and r[1].strip():
            leg[r[0]] = r[1].strip()
    if not leg:
        die("Einbauort-Legende leer — Position-Code-Tabelle nicht gefunden.")
    return leg


def pos_label(code, leg):
    """Label für einen Positionscode: direkt aus Legende, sonst Bitmasken-Dekomposition
    in atomare Zweierpotenz-Codes (Join absteigend, wie in der VDI-Tabelle:
    528 = 512+16 -> 'unten vorne')."""
    if code in leg:
        return leg[code]
    atoms = [a for a in sorted(leg, reverse=True)
             if a > 0 and (a & (a - 1)) == 0]          # Zweierpotenzen der Legende
    rest, parts = code, []
    for a in atoms:
        if rest & a:
            parts.append(leg[a])
            rest &= ~a
    if rest or not parts:
        die(f"Positionscode {code} nicht in Legende und nicht dekomponierbar (Rest {rest}).")
    return " ".join(parts)


def parse_positionen(spec):
    """'10610:16,17;10675:65' oder Datei (TSV/CSV, Spalten id,code) -> [(id, code), ...]"""
    pairs = []
    p = Path(spec)
    if p.exists():
        import re
        for ln, line in enumerate(p.read_text(encoding="utf-8").splitlines(), 1):
            line = line.strip()
            if not line or line.lower().startswith(("id", "#")):
                continue
            toks = [t for t in re.split(r"[\t;,]", line) if t.strip()]
            if len(toks) < 2:
                die(f"--positionen Datei Zeile {ln}: erwarte 'id<TAB>code', bekam {line!r}")
            pairs.append((int(toks[0]), int(toks[1])))
    else:
        for block in spec.split(";"):
            if not block.strip():
                continue
            if ":" not in block:
                die(f"--positionen Spec-Block ohne ':': {block!r}")
            i, codes = block.split(":", 1)
            for c in codes.split(","):
                pairs.append((int(i.strip()), int(c.strip())))
    seen, out = set(), []
    for pr in pairs:
        if pr not in seen:
            seen.add(pr)
            out.append(pr)
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input")
    ap.add_argument("-o", "--output", default=None)
    ap.add_argument("--stammliste", default=None)
    ap.add_argument("--base-uri", default="http://vdi.de/kfz#")
    ap.add_argument("--ontology-uri", default=None)
    ap.add_argument("--partof", choices=("restriction", "direct"), default="restriction")
    ap.add_argument("--vdi-original", default=None,
                    help="Original-VDI-xlsx mit Blatt 'Einbauort' (Legende); nötig für --positionen")
    ap.add_argument("--positionen", default=None,
                    help="Spec 'id:c1,c2;id:c3' oder Pfad zu TSV/CSV mit Spalten id,code")
    args = ap.parse_args()
    if args.positionen and not args.vdi_original:
        die("--positionen benötigt --vdi-original (Quelle der Einbauort-Legende).")

    inp = Path(args.input)
    if not inp.exists():
        die(f"Eingabedatei nicht gefunden: {inp}")
    out = Path(args.output) if args.output else inp.with_suffix(".owl")
    base = args.base_uri if args.base_uri.endswith(("#", "/")) else args.base_uri + "#"
    onto_uri = args.ontology_uri or base.rstrip("#/")

    bau, rels, cats = read_gold(inp)
    stamm = read_stammliste(args.stammliste) if args.stammliste else {}

    known = set(bau) | set(cats) | set(HGN)
    anchors = sorted({p for _, _, p in rels} - known)
    if any(a not in stamm for a in anchors) and not args.stammliste:
        # Selbstauflösung: die Eingabedatei kann die Stammliste enthalten (VDI 5.6)
        try:
            stamm = read_stammliste(str(inp))
        except SystemExit:
            pass
    unresolved = [a for a in anchors if a not in stamm]
    if unresolved:
        die(f"{len(unresolved)} referenzierte IDs sind weder im Gold noch in der Stammliste "
            f"auflösbar: {unresolved[:15]}{' …' if len(unresolved) > 15 else ''} "
            f"-> --stammliste <Template.xlsx> angeben.")

    sub_children = {c for c, rel, _ in rels if rel == "subClassOf"}
    cats_ohne_sub = [i for i in cats if i not in sub_children]
    if cats_ohne_sub:
        die(f"Kategorien ohne sub-Parent (im Gold ergänzen): {cats_ohne_sub}")

    sub_edges = [(c, p) for c, rel, p in rels if rel == "subClassOf"]
    part_edges = [(c, p) for c, rel, p in rels if rel == "partOf"]
    cycle_check(sub_edges, "subClassOf")
    cycle_check(part_edges, "partOf")

    # --- Graph bauen ---
    g = Graph()
    g.bind("owl", OWL)
    g.bind("rdfs", RDFS)
    VDI = Namespace(base)
    g.bind("vdi", VDI)
    onto = URIRef(onto_uri)
    g.add((onto, RDF.type, OWL.Ontology))
    g.add((onto, RDFS.comment, Literal(
        f"VDI-Baugruppen-Goldstandard als Source-Ontologie ({len(bau)} Baugruppen, "
        f"{len(cats)} Zwischenkategorien). Generiert aus {inp.name} via vdi2owl.py; "
        f"partOf-Stil: {args.partof}.", lang=LANG)))

    HAUPTGRUPPE = VDI["hauptgruppe"]
    g.add((HAUPTGRUPPE, RDF.type, OWL.AnnotationProperty))
    g.add((HAUPTGRUPPE, RDFS.label, Literal("Hauptgruppe (final)", lang=LANG)))
    PART_OF = VDI["part_of"]
    g.add((PART_OF, RDF.type, OWL.ObjectProperty))
    g.add((PART_OF, RDFS.label, Literal("part_of", lang="en")))
    g.add((PART_OF, RDFS.label, Literal("ist Teil von", lang=LANG)))

    def uri(i):
        return VDI[f"v{i}"]

    def hguri(n):
        return VDI[f"hg{n}"]

    def target(p):
        return hguri(p) if p in HGN else uri(p)

    for n, name in HGN.items():
        g.add((hguri(n), RDF.type, OWL.Class))
        g.add((hguri(n), RDFS.label, Literal(name, lang=LANG)))
        g.add((hguri(n), RDFS.subClassOf, OWL.Thing))

    def declare(i, label, comment=None):
        c = uri(i)
        g.add((c, RDF.type, OWL.Class))
        g.add((c, RDFS.label, Literal(label, lang=LANG)))
        if comment:
            g.add((c, RDFS.comment, Literal(comment, lang=LANG)))
        return c

    for i, (label, hg) in bau.items():
        c = declare(i, label)
        g.add((c, HAUPTGRUPPE, Literal(hg, lang=LANG)))
    for i, label in cats.items():
        declare(i, label, "Zwischenkategorie aus Adjudikation")
    n_anchor_sub = 0
    for i in anchors:
        label, hg = stamm[i]
        declare(i, label, "Referenzierter Anker außerhalb des Gold-Scopes")
        n = hg_num(hg) if hg else None
        if n:
            g.add((uri(i), RDFS.subClassOf, hguri(n)))
            n_anchor_sub += 1

    for c, rel, p in rels:
        child = uri(c)
        if rel == "subClassOf":
            g.add((child, RDFS.subClassOf, target(p)))
        else:
            if args.partof == "restriction":
                b = BNode()
                g.add((child, RDFS.subClassOf, b))
                g.add((b, RDF.type, OWL.Restriction))
                g.add((b, OWL.onProperty, PART_OF))
                g.add((b, OWL.someValuesFrom, target(p)))
            else:
                g.add((child, PART_OF, target(p)))

    n_hg_anchor = 0
    for i, (label, hg) in bau.items():
        if i not in sub_children:
            g.add((uri(i), RDFS.subClassOf, hguri(hg_num(hg))))
            n_hg_anchor += 1

    n_pos = 0
    if args.positionen:
        leg = read_einbauort_legende(args.vdi_original)
        pairs = parse_positionen(args.positionen)
        bad = sorted({i for i, _ in pairs} - set(bau))
        if bad:
            die(f"--positionen referenziert IDs außerhalb von Gold_Baugruppen: {bad}")
        EINBAUPOS = VDI["einbauposition"]
        g.add((EINBAUPOS, RDF.type, OWL.AnnotationProperty))
        g.add((EINBAUPOS, RDFS.label, Literal("Einbauposition (Code)", lang=LANG)))
        for i, code in pairs:
            base_label, hg = bau[i]
            c = VDI[f"v{i}_{code}"]
            g.add((c, RDF.type, OWL.Class))
            g.add((c, RDFS.label, Literal(f"{base_label}, {pos_label(code, leg)}", lang=LANG)))
            g.add((c, RDFS.subClassOf, uri(i)))
            g.add((c, EINBAUPOS, Literal(str(code))))
            g.add((c, HAUPTGRUPPE, Literal(hg, lang=LANG)))
            n_pos += 1

    g.serialize(destination=str(out), format="pretty-xml")
    print(f"OK: {out}")
    print(f"  Klassen: {len(bau)} Baugruppen + {len(cats)} Kategorien + {len(anchors)} Anker + {len(HGN)} Hauptgruppen")
    print(f"  Kanten : subClassOf {len(sub_edges)} + HG-Verankerungen {n_hg_anchor} + Anker-Verankerungen {n_anchor_sub}")
    print(f"           partOf {len(part_edges)} (Stil: {args.partof})")
    if n_pos:
        print(f"  Positionsklassen: {n_pos} (v{{ID}}_{{Code}}, additiv unter ihren Baugruppen)")


if __name__ == "__main__":
    main()
